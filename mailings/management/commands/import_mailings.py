from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from mailings.models import MailingRecord
from mailings.services import send_email

logger = logging.getLogger(__name__)


REQUIRED_COLUMNS = ["external_id", "user_id", "email", "subject", "message"]


@dataclass
class ImportStats:
    processed_rows: int = 0
    created_records: int = 0
    skipped_records: int = 0
    error_rows: int = 0


class Command(BaseCommand):
    help = "Импорт рассылок из XLSX-файла с последующей отправкой писем."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "xlsx_path",
            type=str,
            help="Путь к XLSX-файлу с данными рассылок",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"Файл {path} не найден.")

        stats = ImportStats()

        self.stdout.write(f"Начинаю обработку файла: {path}")
        workbook = load_workbook(filename=path, read_only=True)
        sheet = workbook.active

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise CommandError("Файл пустой, заголовки не найдены.")

        header_to_index = self._build_header_index(header_row)

        for row in rows_iter:
            stats.processed_rows += 1
            try:
                row_data = self._extract_row_data(row, header_to_index)
            except ValueError as exc:
                stats.error_rows += 1
                logger.warning("Строка %s пропущена из‑за ошибки: %s", stats.processed_rows + 1, exc)
                continue

            external_id = row_data["external_id"]
            try:
                with transaction.atomic():
                    obj, created = MailingRecord.objects.get_or_create(
                        external_id=external_id,
                        defaults={
                            "user_id": row_data["user_id"],
                            "email": row_data["email"],
                            "subject": row_data["subject"],
                            "message": row_data["message"],
                        },
                    )

                    if not created:
                        stats.skipped_records += 1
                        logger.info("Запись с external_id=%s уже существует, пропускаю.", external_id)
                        continue

                    stats.created_records += 1

                    try:
                        send_email(
                            user_id=obj.user_id,
                            email=obj.email,
                            subject=obj.subject,
                            message=obj.message,
                        )
                        obj.status = MailingRecord.Status.SENT
                        obj.error_message = ""
                    except Exception as exc:  # noqa: BLE001
                        obj.status = MailingRecord.Status.FAILED
                        obj.error_message = str(exc)
                        stats.error_rows += 1
                        logger.exception("Ошибка при отправке письма для external_id=%s", external_id)

                    obj.save(update_fields=["status", "error_message", "processed_at"])
            except Exception:  # noqa: BLE001
                stats.error_rows += 1
                logger.exception("Непредвиденная ошибка при обработке строки с external_id=%s", external_id)

        self._print_summary(stats)

    def _build_header_index(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        header_to_index: dict[str, int] = {}
        for idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            header_to_index[str(cell_value).strip()] = idx

        missing = [col for col in REQUIRED_COLUMNS if col not in header_to_index]
        if missing:
            raise CommandError(f"В файле отсутствуют обязательные колонки: {', '.join(missing)}")
        return header_to_index

    def _extract_row_data(self, row: tuple[Any, ...], header_to_index: dict[str, int]) -> dict[str, str]:
        data: dict[str, str] = {}
        for col in REQUIRED_COLUMNS:
            idx = header_to_index[col]
            value = row[idx] if idx < len(row) else None
            if value is None or str(value).strip() == "":
                raise ValueError(f"Пустое значение в обязательной колонке {col!r}")
            data[col] = str(value).strip()
        return data

    def _print_summary(self, stats: ImportStats) -> None:
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Импорт завершён."))
        self.stdout.write(f"Количество обработанных строк: {stats.processed_rows}")
        self.stdout.write(f"Количество созданных записей: {stats.created_records}")
        self.stdout.write(f"Количество пропущенных записей: {stats.skipped_records}")
        self.stdout.write(f"Количество ошибочных строк: {stats.error_rows}")

